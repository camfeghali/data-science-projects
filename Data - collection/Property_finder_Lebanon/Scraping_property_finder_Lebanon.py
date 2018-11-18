import requests, bs4
import time
import openpyxl
import pprint
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('Property_Finder.xlsx')
sheet = wb['Sheet1']

addLinks=[]
addInfo=[]
facts_soupObjects=[]
seller_facts_soupObjects=[]
company_list=[]
location_list=[]
lat_list=[]
lon_list=[]
amenity_soupObject=[]
Add_dictionnary={}

page_num=1
add_number=1
location_number=1

property_number=1

while page_num<=381:
    res_finder=requests.get("https://www.propertyfinder.com.lb/en/search?c=1&ob=mr&page="+str(page_num))
    res_finder.raise_for_status()
    baseUrl="https://www.propertyfinder.com.lb"
    soup_finder = bs4.BeautifulSoup(res_finder.text)
    time.sleep(0.5)

    ##GET PROPERTY ADDLINK
    for prop in soup_finder.find_all("div", class_="cardlist_item"):
        link=prop.a['href']
        linkUrl=baseUrl+link
        addLinks.append(linkUrl)
        sheet.cell(add_number,1).value = linkUrl
        add_number=add_number+1


    ##GET PROPERTY LOCATION
    for location in soup_finder.find_all("p", class_="card_location"):
        location=str(location).replace(" ","").replace("\n","").split("svg>")[1].replace("</p>","")
        location_list.append(location)
        sheet.cell(location_number,2).value = location
        location_number=location_number+1

    page_num=page_num+1
    print("Page number is: "+ str(page_num))
#     print(page_num)




for cellObj in sheet['A']: ##here the A represents the column name
    try:
        link = cellObj.value
        ress = requests.get(link)
        time.sleep(0.5)
        ress.raise_for_status()

        ##PROPERTY INFO##
        soup_finder = bs4.BeautifulSoup(ress.text)
        propFacts=soup_finder.find("div", class_="facts_container")
        propFacts1 = bs4.BeautifulSoup(propFacts.text)
        facts_soupObjects.append(propFacts1)

        ##AMENITIES INFO
        propAmenities=soup_finder.find("div", class_="amenities_container")
        amenity_soupObject.append(propAmenities)

        ##SELLER INFO##
        sellerFacts = soup_finder.find("div", class_="agentinfo_detail")
        sellerFacts1 = bs4.BeautifulSoup(sellerFacts.text)
        seller_facts_soupObjects.append(sellerFacts1)

        ##GEOLCATION DATA##
        lat=soup_finder.getText().split("latitude")[1].split(",")[0].split(":")[1]
        lon=soup_finder.getText().split("longitude")[1].split(",")[0].split(":")[1].split("}")[0]
        lat_list.append(lat)
        lon_list.append(lon)
        property_number=property_number+1

        print("property being scraped is: "+str(property_number))
    except:'AttributeError'

##PARSE AND INPUT AMENITIES

for i in range (0,len(amenity_soupObject)):

    if type(amenity_soupObject[i]) is bs4.element.Tag:

        amenities=str(amenity_soupObject[i].getText()).replace(" ","/").replace("\n"," ").split("Listed")[0].replace(" ","").replace("/","")
        amenities=amenities.replace("Amenities","").replace("Balcony","Balcony,").replace("BuiltinWardrobes","Built in Wardrobes,").replace("CentralAC","Central AC,").replace("CoveredParking","Covered Parking,")
        amenities = amenities.replace("KitchenAppliances","Kitchen Appliances,").replace("MaidsRoom","Maids Room,").replace("PetsAllowed","Pets Allowed,").replace("SecurityStudy","Security study,").replace("ViewofLandmark","View of Landmark,")
        amenities = amenities.replace("Concierge","Concierge,").replace("PrivateGarden","Private Garden").replace("Security","Security,").replace("SharedGym","Shared Gym,")
        amenities = amenities.replace("SharedPool","Shared pool,").replace("Networked","Networked,").replace("ViewofWater","View of Water,")
        amenities = amenities.replace("Walk-inCloset","Walk-in-Closet,").replace("LobbyinBuildin","Lobby in Building").replace("Study","Study,").replace("Children'sPlayArea","children play area,").replace("Children'sPool","Children pool,")
        sheet.cell(i+1,13).value = amenities
    else:
        sheet.cell(i+1,13).value="no amenities here"

##PARSE AND ADD LATITUDE AND LONGITUDE

for i in range (0,len(lat_list)):

    lat=lat_list[i]
    lon=lon_list[i]
    sheet.cell(i+1,11).value = lat
    sheet.cell(i+1,12).value = lon

##PARSE AND ADD SELLER INFO

for i in range (0,len(seller_facts_soupObjects)):

    agent=str(seller_facts_soupObjects[i]).replace("\n","").split("Agent:")[1].split(":")[0].split("Viewallourproperties")[0].replace("Company","")
    company=str(seller_facts_soupObjects[i]).replace("\n","").split("Company:")[1].split("Viewallourproperties")[0].split("View")[0]
    sheet.cell(i+1,9).value = agent
    sheet.cell(i+1,8).value = company


##PARSE AND ADD PROPERTY INFO
add_numba=1
for soupObject in facts_soupObjects:

    ######OBJECT TEXT#####
    text=soupObject.getText()
    text=text.replace(" ","").replace("\n","").replace("FactsPrice","FactsPrice//").replace("Type","//Type//").replace("Reference","//Reference//").replace("Bedrooms","//Bedrooms//").replace("Bathrooms","//Bathrooms//").replace("Area","//Area//")
#     print(text)

    type_cat=text.split("USD")[1][2:6]

    values = text.split("//")[1::2]
    categories = text.split("//")[::2]

#     print(categories)
#     print(values)



##PARSE AND ADD FACTS
    for i in range(0,len(categories)):
#         print(categories[i] + " " +values[i])

        if categories[i] == "FactsPrice":
            sheet.cell(add_numba,4).value = values[i]
        if categories[i] == "Type":
            sheet.cell(add_numba,3).value = values[i]
        if categories[i] == "Reference":
            sheet.cell(add_numba,10).value = values[i]
        if categories[i] == "Bedrooms":
            sheet.cell(add_numba,5).value = values[i]
        if categories[i] == "Bathrooms":
            sheet.cell(add_numba,6).value = values[i]
        if categories[i] == "Area":
            sheet.cell(add_numba,7).value = values[i]

    add_numba=add_numba+1

    ######################

wb.save('Property_Finder_LEBANON.xlsx')



print("facts")
print(len(seller_facts_soupObjects))
print("amenities")
print(len(amenity_soupObject))
print("Seller Info")
print(len(seller_facts_soupObjects))
print("Lat")
print (len(lat_list))
print("Lon")
print (len(lon_list))
print("Links")
print(len(addLinks))